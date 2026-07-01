from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


CODE_KEYS = (
    "account_code",
    "hesap_kodu",
    "hesap kodu",
    "kod",
    "code",
    "raw_account_code",
)
NAME_KEYS = (
    "account_name",
    "hesap_adi",
    "hesap adı",
    "hesap adi",
    "name",
    "description",
)
TAX_ID_KEYS = ("tax_id", "vkn", "tckn", "vergi_no", "vergi no")
TAX_OFFICE_KEYS = ("tax_office", "vergi_dairesi", "vergi dairesi")
IBAN_KEYS = ("iban", "banka_iban", "banka iban", "counterparty_iban", "karsi_iban")
DETAIL_KEYS = ("is_detail_account", "detay e/h", "detay", "detay hesap")
VAT_RATE_HINTS = {"001": "1", "008": "8", "010": "10", "018": "18", "020": "20"}
USAGE_KEYWORDS = {
    "cihaz": ("cihaz", "makine", "malzeme"),
    "pil": ("pil",),
    "kalip": ("kalip", "kalib"),
    "elektrik": ("elektrik",),
    "su": (" su ", "su gider", "su fatur"),
    "kira": ("kira",),
    "arac_kiralama": ("arac kiralama",),
    "kargo": ("kargo", "nakliye"),
    "akaryakit": ("akaryakit", "yakit", "benzin", "motorin"),
    "hgs": ("hgs",),
    "musavirlik": ("musavir", "muhasebe"),
    "bakim": ("bakim", "tamir", "onarim"),
    "satis": ("satis", "satislar"),
    "tevkifat": ("tevkifat",),
    "internet": ("internet",),
    "yazilim": ("yazilim", "e-fatura", "efatura"),
    "guvenlik": ("guvenlik", "güvenlik", "security"),
}


@dataclass(frozen=True)
class ChartAccount:
    raw_account_code: str
    normalized_account_code: str
    account_name: str
    is_detail_account: bool | None = None
    tax_id: str | None = None
    tax_office: str | None = None
    iban: str | None = None

    @property
    def is_counterparty_candidate(self) -> bool:
        return self.normalized_account_code.startswith(("120", "320"))

    @property
    def counterparty_type(self) -> str | None:
        if self.normalized_account_code.startswith("120"):
            return "customer"
        if self.normalized_account_code.startswith("320"):
            return "supplier"
        return None

    @property
    def account_family(self) -> str:
        match = re.match(r"^(\d{3})", self.normalized_account_code)
        return match.group(1) if match else ""

    @property
    def code_depth(self) -> int:
        return len([part for part in self.normalized_account_code.split(".") if part])

    @property
    def vat_rate_hint(self) -> str:
        code_parts = [part for part in self.normalized_account_code.split(".") if part]
        for part in reversed(code_parts):
            rate = VAT_RATE_HINTS.get(part.zfill(3))
            if rate:
                return rate
        normalized_name = normalize_text(self.account_name)
        match = re.search(r"(?:%|yuzde\s*)(0?1|0?8|10|18|20)\b", normalized_name)
        if match:
            return str(int(match.group(1)))
        return ""

    @property
    def usage_tags(self) -> tuple[str, ...]:
        normalized_name = f" {normalize_text(self.account_name)} "
        tags: list[str] = []
        for tag, needles in USAGE_KEYWORDS.items():
            if any(needle in normalized_name for needle in needles):
                tags.append(tag)
        return tuple(tags)


def normalize_account_code(value: str) -> str:
    compact = value.strip().replace(",", ".")
    compact = re.sub(r"[\s\-]+", ".", compact)
    compact = re.sub(r"[^0-9A-Za-z.]", "", compact)
    compact = re.sub(r"\.+", ".", compact).strip(".")
    return compact


def normalize_text(value: str) -> str:
    replacements = {
        "\u0131": "i",
        "\u0130": "i",
        "I": "i",
        "\u015f": "s",
        "\u015e": "s",
        "\u011f": "g",
        "\u011e": "g",
        "\u00fc": "u",
        "\u00dc": "u",
        "\u00f6": "o",
        "\u00d6": "o",
        "\u00e7": "c",
        "\u00c7": "c",
        "\u00e2": "a",
        "\u00c2": "a",
        "Ä±": "i",
        "Ä°": "i",
        "ÄŸ": "g",
        "Ä": "g",
        "Ã¼": "u",
        "Ãœ": "u",
        "ÅŸ": "s",
        "Å": "s",
        "Ã¶": "o",
        "Ã–": "o",
        "Ã§": "c",
        "Ã‡": "c",
    }
    result = value
    for source, target in replacements.items():
        result = result.replace(source, target)
    decomposed = unicodedata.normalize("NFKD", result)
    asciiish = "".join(character for character in decomposed if not unicodedata.combining(character))
    return re.sub(r"\s+", " ", asciiish.lower()).strip()


def normalize_iban(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z]", "", value).upper()


def _normalized_header(value: str) -> str:
    return normalize_text(value).replace("_", " ")


def _first_value(row: dict[str, str], keys: Iterable[str]) -> str:
    normalized = {_normalized_header(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(_normalized_header(key))
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _is_child(parent: str, child: str) -> bool:
    if parent == child or not child.startswith(parent):
        return False
    remainder = child[len(parent) :]
    if remainder.startswith((".", "-")):
        return True
    return parent.isdigit() and child.isdigit() and len(child) > len(parent)


def parse_detail_flag(value: str) -> bool | None:
    normalized = _normalized_header(value)
    if normalized in {"evet", "e", "yes", "true", "1"}:
        return True
    if normalized in {"hayir", "h", "no", "false", "0"}:
        return False
    return None


def mark_detail_accounts(accounts: list[ChartAccount]) -> list[ChartAccount]:
    codes = [account.normalized_account_code for account in accounts]
    result: list[ChartAccount] = []
    for account in accounts:
        has_child = any(_is_child(account.normalized_account_code, code) for code in codes)
        explicit_detail = account.is_detail_account
        result.append(
            ChartAccount(
                raw_account_code=account.raw_account_code,
                normalized_account_code=account.normalized_account_code,
                account_name=account.account_name,
                is_detail_account=explicit_detail if explicit_detail is not None else not has_child,
                tax_id=account.tax_id,
                tax_office=account.tax_office,
                iban=account.iban,
            )
        )
    return result


def parse_chart_accounts_csv(path: Path) -> list[ChartAccount]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        accounts = []
        for row in reader:
            raw_code = _first_value(row, CODE_KEYS)
            if not raw_code:
                continue
            name = _first_value(row, NAME_KEYS) or raw_code
            accounts.append(
                ChartAccount(
                    raw_account_code=raw_code,
                    normalized_account_code=normalize_account_code(raw_code),
                    account_name=name,
                    is_detail_account=parse_detail_flag(_first_value(row, DETAIL_KEYS)),
                    tax_id=_first_value(row, TAX_ID_KEYS) or None,
                    tax_office=_first_value(row, TAX_OFFICE_KEYS) or None,
                    iban=normalize_iban(_first_value(row, IBAN_KEYS)) or None,
                )
            )
    return mark_detail_accounts(accounts)


def parse_chart_accounts_xlsx(path: Path) -> list[ChartAccount]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX parsing requires openpyxl. Install backend requirements first.") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "") for value in rows[0]]
    accounts = []
    for values in rows[1:]:
        row = {headers[index]: str(value or "") for index, value in enumerate(values)}
        raw_code = _first_value(row, CODE_KEYS)
        if not raw_code:
            continue
        name = _first_value(row, NAME_KEYS) or raw_code
        accounts.append(
            ChartAccount(
                raw_account_code=raw_code,
                normalized_account_code=normalize_account_code(raw_code),
                account_name=name,
                is_detail_account=parse_detail_flag(_first_value(row, DETAIL_KEYS)),
                tax_id=_first_value(row, TAX_ID_KEYS) or None,
                tax_office=_first_value(row, TAX_OFFICE_KEYS) or None,
                iban=normalize_iban(_first_value(row, IBAN_KEYS)) or None,
            )
        )
    return mark_detail_accounts(accounts)


def parse_chart_accounts(path: Path | str) -> list[ChartAccount]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return parse_chart_accounts_csv(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_chart_accounts_xlsx(file_path)
    raise ValueError(f"Unsupported chart account format: {suffix}")


def extract_counterparty_candidates(accounts: Iterable[ChartAccount]) -> list[ChartAccount]:
    return [account for account in accounts if account.is_detail_account and account.is_counterparty_candidate]


def validate_vat_accounts(accounts: Iterable[ChartAccount]) -> dict[str, bool]:
    codes = {account.normalized_account_code for account in accounts}
    return {
        "has_purchase_vat_191": any(code.startswith("191") for code in codes),
        "has_sales_vat_391": any(code.startswith("391") for code in codes),
    }


def _detail_accounts(accounts: Iterable[ChartAccount], prefixes: tuple[str, ...]) -> list[ChartAccount]:
    return [
        account
        for account in accounts
        if account.is_detail_account and account.normalized_account_code.startswith(prefixes)
    ]


def _deepest(accounts: Iterable[ChartAccount]) -> ChartAccount | None:
    candidates = list(accounts)
    if not candidates:
        return None
    return max(candidates, key=lambda account: (account.code_depth, len(account.normalized_account_code)))


def _rate_text(vat_rate: str | int | None) -> str:
    raw = str(vat_rate or "").strip()
    return str(int(raw)) if raw.isdigit() else ""


def select_vat_account(accounts: Iterable[ChartAccount], account_family: str, vat_rate: str | int | None) -> ChartAccount | None:
    family = str(account_family)
    normalized_rate = _rate_text(vat_rate)
    candidates = _detail_accounts(accounts, (family,))
    if normalized_rate:
        selected = _deepest(account for account in candidates if account.vat_rate_hint == normalized_rate)
        if selected:
            return selected
    return _deepest(candidates)


def select_revenue_account(accounts: Iterable[ChartAccount], vat_rate: str | int | None = None) -> ChartAccount | None:
    return select_vat_like_revenue_account(accounts, "600", vat_rate)


def select_vat_like_revenue_account(
    accounts: Iterable[ChartAccount],
    family: str,
    vat_rate: str | int | None = None,
) -> ChartAccount | None:
    normalized_rate = _rate_text(vat_rate)
    candidates = _detail_accounts(accounts, (family,))
    if normalized_rate:
        selected = _deepest(account for account in candidates if account.vat_rate_hint == normalized_rate)
        if selected:
            return selected
    return _deepest(candidates)


def select_usage_account(
    accounts: Iterable[ChartAccount],
    line_hint: str,
    direction: str,
    *,
    account_treatment: str = "",
) -> ChartAccount | None:
    normalized_hint = f" {normalize_text(line_hint)} "
    desired_tags = {
        tag
        for tag, needles in USAGE_KEYWORDS.items()
        if any(needle in normalized_hint for needle in needles)
    }
    treatment = normalize_text(account_treatment)
    if "mal" in normalized_hint or "stok" in treatment or desired_tags.intersection({"cihaz", "pil", "kalip"}):
        preferred_prefixes = ("153", "740", "760", "770")
    elif direction == "sales":
        preferred_prefixes = ("600",)
    else:
        preferred_prefixes = ("770", "760", "740", "153")
    candidates = _detail_accounts(accounts, preferred_prefixes)
    if not candidates:
        return None

    if not desired_tags:
        generic_candidates = [
            account
            for account in candidates
            if any(needle in normalize_text(account.account_name) for needle in ("genel", "disaridan", "diger"))
        ]
        return min(generic_candidates or candidates, key=lambda account: (account.code_depth, len(account.normalized_account_code), account.normalized_account_code))

    def score(account: ChartAccount) -> tuple[int, int, int, str]:
        tags = set(account.usage_tags)
        tag_score = len(desired_tags.intersection(tags)) * 20
        if account.account_family == "153" and desired_tags.intersection({"cihaz", "pil", "kalip"}):
            tag_score += 15
        if account.account_family in {"770", "760", "740"} and desired_tags and desired_tags.isdisjoint({"cihaz", "pil", "kalip"}):
            tag_score += 5
        return (tag_score, account.code_depth, len(account.normalized_account_code), account.normalized_account_code)

    return max(candidates, key=score)
