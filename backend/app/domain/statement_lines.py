from __future__ import annotations

import csv
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
import io
from pathlib import Path
import re
from typing import Iterable
import unicodedata

from app.domain.chart_accounts import ChartAccount, normalize_account_code
from app.domain.counterparty_matching import match_counterparty
from app.domain.statement_rule_engine import classify_statement_transaction


@dataclass(frozen=True)
class StatementLine:
    line_no: int
    transaction_date: str
    description: str
    amount: str
    direction: str
    balance_after: str = ""
    counterparty_name: str = ""
    tax_id: str = ""
    iban: str = ""
    suggested_account_code: str = ""
    transaction_type: str = "unknown"
    confidence: int = 35
    risk_flags: tuple[str, ...] = ("statement_review_required",)
    review_reason: str = "no_statement_rule_matched"
    counterparty_match_code: str = ""
    counterparty_match_name: str = ""
    counterparty_match_confidence: int = 0
    counterparty_match_reason: str = "not_assessed"


HEADER_ALIASES = {
    "transaction_date": {
        "transaction_date",
        "date",
        "tarih",
        "tarih_saat",
        "hareket_tarih",
        "hareket_tarihi",
        "islem_tarihi",
        "işlem_tarihi",
    },
    "description": {"description", "aciklama", "açıklama", "islem_aciklama", "işlem_açıklama"},
    "amount": {"amount", "tutar", "islem_tutari", "işlem_tutarı"},
    "debit": {"debit", "borc", "borç", "borc_tutari", "borç_tutarı", "cikis", "çıkış", "cikis_tutari", "çıkış_tutarı"},
    "credit": {"credit", "alacak", "alacak_tutari", "giris", "giriş", "giris_tutari", "giriş_tutarı"},
    "direction": {"direction", "yon", "yön", "borc_alacak", "borç_alacak"},
    "balance_after": {"balance_after", "bakiye", "son_bakiye"},
    "counterparty_name": {"counterparty_name", "cari_unvan", "unvan", "firma", "alici_satici", "karsi_hesap"},
    "tax_id": {"tax_id", "vkn", "tckn", "vergi_no", "vergi_numarasi"},
    "iban": {"iban", "karsi_iban", "counterparty_iban"},
    "suggested_account_code": {"suggested_account_code", "hesap_kodu", "account_code"},
}

IDENTITY_STATEMENT_MAPPING = {
    "transaction_date": "transaction_date",
    "description": "description",
    "amount": "amount",
    "debit": "debit",
    "credit": "credit",
    "direction": "direction",
    "balance_after": "balance_after",
    "counterparty_name": "counterparty_name",
    "tax_id": "tax_id",
    "iban": "iban",
    "suggested_account_code": "suggested_account_code",
}

TEXT_DECODINGS = ("utf-8-sig", "utf-8", "cp1254", "iso-8859-9", "latin-1")
OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0"
GENERIC_COUNTERPARTY_ACCOUNTS = {"120", "320"}
COUNTERPARTY_MATCH_TYPES = {"unknown", "bank_transfer_in", "bank_transfer_out"}


class _TableRowsParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "tr":
            self._current_row = []
        if tag.lower() in {"td", "th"} and self._current_row is not None:
            self._current_cell = []

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        normalized = tag.lower()
        if normalized in {"td", "th"} and self._current_row is not None and self._current_cell is not None:
            self._current_row.append(" ".join(part.strip() for part in self._current_cell if part.strip()))
            self._current_cell = None
        if normalized == "tr" and self._current_row is not None:
            if any(cell.strip() for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None
            self._current_cell = None


def _normalize(value: str) -> str:
    replacements = {
        "ı": "i",
        "İ": "i",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ş": "s",
        "Ş": "s",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c",
    }
    result = value.strip()
    for source, target in replacements.items():
        result = result.replace(source, target)
    result = unicodedata.normalize("NFKD", result)
    result = "".join(character for character in result if not unicodedata.combining(character))
    result = result.lower()
    result = re.sub(r"[^a-z0-9]+", "_", result).strip("_")
    return result


def _column_map(headers: Iterable[str]) -> dict[str, str]:
    normalized_headers = {header: _normalize(header) for header in headers}
    mapped: dict[str, str] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for header, normalized in normalized_headers.items():
            if normalized in aliases:
                mapped[canonical] = header
                break
    return mapped


def _parse_decimal(value: str) -> Decimal | None:
    compact = value.strip().replace(" ", "").replace("\xa0", "")
    if not compact:
        return None
    if compact.startswith("(") and compact.endswith(")"):
        compact = f"-{compact[1:-1]}"
    compact = re.sub(r"(?i)(try|tl|₺)", "", compact)
    compact = re.sub(r"[^0-9,.\-]", "", compact)
    if not compact or compact in {"-", ".", ","}:
        return None
    if "," in compact and "." in compact:
        compact = compact.replace(".", "").replace(",", ".")
    elif "," in compact:
        compact = compact.replace(",", ".")
    try:
        return Decimal(compact)
    except InvalidOperation:
        return None


def _format_decimal(value: Decimal | None) -> str:
    return f"{value:.2f}" if value is not None else ""


def _is_nonzero(value: Decimal | None) -> bool:
    return value is not None and value != 0


def _string_value(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")


def _amount_and_direction_from_values(
    *,
    amount: str,
    debit: str = "",
    credit: str = "",
    direction: str = "",
) -> tuple[Decimal | None, str]:
    parsed_amount = _parse_decimal(amount)
    debit_amount = _parse_decimal(debit)
    credit_amount = _parse_decimal(credit)
    if _is_nonzero(debit_amount):
        return abs(debit_amount), "out"
    if _is_nonzero(credit_amount):
        return abs(credit_amount), "in"
    return parsed_amount, _infer_direction(parsed_amount, direction)


def _infer_direction(amount: Decimal | None, raw_direction: str) -> str:
    direction = _normalize(raw_direction)
    if direction in {"out", "cikis", "borc", "debit", "odeme"}:
        return "out"
    if direction in {"in", "giris", "alacak", "credit", "tahsilat"}:
        return "in"
    if amount is not None and amount < 0:
        return "out"
    if amount is not None and amount > 0:
        return "in"
    return ""


def _line_from_row(line_no: int, row: dict[str, object], mapping: dict[str, str]) -> StatementLine:
    def value(key: str) -> str:
        column = mapping.get(key, "")
        return _string_value(row.get(column, "")).strip()

    amount, direction = _amount_and_direction_from_values(
        amount=value("amount"),
        debit=value("debit"),
        credit=value("credit"),
        direction=value("direction"),
    )
    decision = classify_statement_transaction(
        description=value("description"),
        direction=direction,
        amount=amount,
        suggested_account_code=value("suggested_account_code"),
    )
    return StatementLine(
        line_no=line_no,
        transaction_date=value("transaction_date"),
        description=value("description"),
        amount=_format_decimal(abs(amount) if amount is not None else None),
        direction=direction,
        balance_after=value("balance_after"),
        counterparty_name=value("counterparty_name"),
        tax_id=value("tax_id"),
        iban=value("iban"),
        suggested_account_code=decision.suggested_account_code,
        transaction_type=decision.transaction_type,
        confidence=decision.confidence,
        risk_flags=decision.risk_flags,
        review_reason=decision.review_reason,
    )


def enrich_statement_lines_with_counterparties(
    lines: tuple[StatementLine, ...],
    accounts: list[ChartAccount],
    learning_events: Iterable[dict[str, object]] = (),
) -> tuple[StatementLine, ...]:
    if not accounts:
        return tuple(
            _add_counterparty_missing_flag(line)
            if line.transaction_type == "unknown" and not line.suggested_account_code
            else line
            for line in lines
        )
    events = tuple(learning_events)
    return tuple(_enrich_statement_line_with_counterparty(line, accounts, events) for line in lines)


def _enrich_statement_line_with_counterparty(
    line: StatementLine,
    accounts: list[ChartAccount],
    learning_events: tuple[dict[str, object], ...],
) -> StatementLine:
    if line.transaction_type in {"tax_payment", "sgk_payment"}:
        return line
    name_hint = line.counterparty_name or line.description
    match = match_counterparty(
        accounts,
        tax_ids=tuple(tax_id for tax_id in (line.tax_id,) if tax_id),
        ibans=tuple(iban for iban in (line.iban,) if iban),
        name_hint=name_hint,
    )
    learned = _learning_counterparty_match(line, accounts, learning_events)
    if learned is not None and (not match.account_code or match.requires_review):
        return learned
    if not match.account_code:
        return _add_counterparty_missing_flag(line) if line.transaction_type == "unknown" else line

    flags = tuple(flag for flag in line.risk_flags if flag != "statement_review_required")
    if match.requires_review:
        flags = tuple(dict.fromkeys((*flags, "counterparty_match_review_required")))
    transaction_type = line.transaction_type
    if transaction_type in COUNTERPARTY_MATCH_TYPES:
        transaction_type = "counterparty_collection" if line.direction == "in" else "counterparty_payment"
    suggested_account_code = line.suggested_account_code
    if not suggested_account_code or suggested_account_code in GENERIC_COUNTERPARTY_ACCOUNTS:
        suggested_account_code = match.account_code
    return replace(
        line,
        suggested_account_code=suggested_account_code,
        transaction_type=transaction_type,
        confidence=max(line.confidence, match.confidence),
        risk_flags=flags,
        counterparty_match_code=match.account_code,
        counterparty_match_name=match.account_name,
        counterparty_match_confidence=match.confidence,
        counterparty_match_reason=match.match_reason,
    )


def _learning_counterparty_match(
    line: StatementLine,
    accounts: list[ChartAccount],
    learning_events: tuple[dict[str, object], ...],
) -> StatementLine | None:
    normalized_text = _normalize(f"{line.description} {line.counterparty_name}")
    if not normalized_text:
        return None
    accounts_by_code = {account.normalized_account_code: account for account in accounts if account.is_detail_account}
    for event in reversed(learning_events):
        account_code = normalize_account_code(str(event.get("corrected_counterparty_code") or ""))
        account = accounts_by_code.get(account_code)
        if account is None:
            continue
        hints = (
            str(event.get("category") or ""),
            str(event.get("reason") or ""),
            str(event.get("document_ref") or ""),
        )
        if not any(_learning_hint_matches(normalized_text, hint) for hint in hints):
            continue
        automation_candidate = bool(event.get("automation_candidate"))
        flags = tuple(flag for flag in line.risk_flags if flag != "statement_review_required")
        if not automation_candidate:
            flags = tuple(dict.fromkeys((*flags, "learning_rule_review_required")))
        transaction_type = line.transaction_type
        if transaction_type in COUNTERPARTY_MATCH_TYPES:
            transaction_type = "counterparty_collection" if line.direction == "in" else "counterparty_payment"
        suggested_account_code = line.suggested_account_code
        if not suggested_account_code or suggested_account_code in GENERIC_COUNTERPARTY_ACCOUNTS:
            suggested_account_code = account.normalized_account_code
        return replace(
            line,
            suggested_account_code=suggested_account_code,
            transaction_type=transaction_type,
            confidence=max(line.confidence, 90 if automation_candidate else 76),
            risk_flags=flags,
            counterparty_match_code=account.normalized_account_code,
            counterparty_match_name=account.account_name,
            counterparty_match_confidence=90 if automation_candidate else 76,
            counterparty_match_reason="learning_event",
        )
    return None


def _learning_hint_matches(normalized_text: str, hint: str) -> bool:
    normalized_hint = _normalize(hint)
    return len(normalized_hint) >= 4 and normalized_hint in normalized_text


def _add_counterparty_missing_flag(line: StatementLine) -> StatementLine:
    return replace(
        line,
        risk_flags=tuple(dict.fromkeys((*line.risk_flags, "counterparty_not_found"))),
        counterparty_match_reason="not_found",
    )


def _decode_text_export(content: bytes) -> str:
    for encoding in TEXT_DECODINGS:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _grid_from_html_table(text: str) -> list[list[object]]:
    if "<table" not in text.lower() and "<tr" not in text.lower():
        return []
    parser = _TableRowsParser()
    parser.feed(text)
    rows = [row for row in parser.rows if any(cell.strip() for cell in row)]
    if len(rows) < 2:
        return []
    return rows


def _unique_headers(values: Iterable[object]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = _string_value(value).strip() or f"__empty_{index}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def _mapping_can_parse_statement(mapping: dict[str, str]) -> bool:
    has_amount = "amount" in mapping or "debit" in mapping or "credit" in mapping
    return "transaction_date" in mapping and "description" in mapping and has_amount


def _row_has_statement_data(row: dict[str, object], mapping: dict[str, str]) -> bool:
    def value(key: str) -> str:
        column = mapping.get(key, "")
        return _string_value(row.get(column, "")).strip()

    amount, direction = _amount_and_direction_from_values(
        amount=value("amount"),
        debit=value("debit"),
        credit=value("credit"),
        direction=value("direction"),
    )
    return bool(value("transaction_date") and value("description") and amount is not None and direction in {"in", "out"})


def _lines_from_row_dicts(rows: list[dict[str, object]], headers: Iterable[str]) -> tuple[StatementLine, ...]:
    mapping = _column_map(headers)
    if not _mapping_can_parse_statement(mapping):
        return ()
    parsed: list[StatementLine] = []
    for row in rows:
        if not _row_has_statement_data(row, mapping):
            continue
        parsed.append(_line_from_row(len(parsed) + 1, row, mapping))
    return tuple(parsed)


def _lines_from_grid(grid: list[list[object]]) -> tuple[StatementLine, ...]:
    best: tuple[int, int, list[str], list[dict[str, object]]] | None = None
    for header_index, header_values in enumerate(grid[:-1]):
        headers = _unique_headers(header_values)
        mapping = _column_map(headers)
        if not _mapping_can_parse_statement(mapping):
            continue
        rows = [
            {header: value for header, value in zip(headers, values, strict=False)}
            for values in grid[header_index + 1 :]
            if any(_string_value(value).strip() for value in values)
        ]
        parseable_count = sum(1 for row in rows if _row_has_statement_data(row, mapping))
        if parseable_count == 0:
            continue
        score = parseable_count * 10 + len(mapping)
        if best is None or score > best[0]:
            best = (score, header_index, headers, rows)
    if best is None:
        return ()
    _, _, headers, rows = best
    return _lines_from_row_dicts(rows, headers)


def _parse_delimited_statement_text(text: str) -> tuple[StatementLine, ...]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        return ()
    delimiter = max(("\t", ";", "|", ","), key=lambda candidate: sum(line.count(candidate) for line in lines[:20]))
    if sum(line.count(delimiter) for line in lines[:20]) == 0:
        return ()
    reader = csv.reader(io.StringIO("\n".join(lines)), delimiter=delimiter)
    return _lines_from_grid([row for row in reader])


def _parse_loose_statement_text(text: str) -> tuple[StatementLine, ...]:
    date_pattern = r"(?P<transaction_date>\d{4}[-/.]\d{1,2}[-/.]\d{1,2}|\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})"
    amount_pattern = r"-?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})|-?\d+(?:[.,]\d{2})"
    direction_pattern = r"in|out|giris|giriş|cikis|çıkış|borc|borç|alacak|debit|credit|odeme|ödeme|tahsilat"
    date_re = re.compile(rf"^{date_pattern}\s+(?P<body>.+?)\s*$", re.IGNORECASE)
    amount_re = re.compile(amount_pattern)
    direction_re = re.compile(rf"\s+(?P<direction>{direction_pattern})\s*$", re.IGNORECASE)
    signed_currency_amount_re = re.compile(rf"(?P<amount>[+-](?:{amount_pattern}))\s*(?:tl|try|₺)?", re.IGNORECASE)
    parsed: list[StatementLine] = []
    for line in (item.strip() for item in text.splitlines() if item.strip()):
        if _normalize(line) in {
            "transaction_date_description_amount_direction",
            "tarih_aciklama_tutar_yon",
            "tarih_aciklama_borc_alacak_bakiye",
        }:
            continue
        match = date_re.match(line)
        if not match:
            continue
        body = match.group("body").strip()
        direction_match = direction_re.search(body)
        direction = direction_match.group("direction") if direction_match else ""
        if direction_match:
            body = body[: direction_match.start()].strip()
        amounts = list(amount_re.finditer(body))
        if not amounts:
            continue
        trailing = [amounts[-1]]
        cursor = amounts[-1].start()
        for amount_match in reversed(amounts[:-1]):
            if body[amount_match.end() : cursor].strip():
                break
            trailing.insert(0, amount_match)
            cursor = amount_match.start()
        description = body[: trailing[0].start()].strip()
        if not description:
            continue
        values: dict[str, str] = {
            "transaction_date": match.group("transaction_date"),
            "description": description,
            "direction": direction,
        }
        if len(trailing) >= 3:
            values.update(
                {
                    "debit": trailing[-3].group(0),
                    "credit": trailing[-2].group(0),
                    "balance_after": trailing[-1].group(0),
                }
            )
        elif len(trailing) == 2 and not direction:
            values.update({"amount": trailing[-2].group(0), "balance_after": trailing[-1].group(0)})
        else:
            values["amount"] = trailing[-1].group(0)
        parsed_amount = _parse_decimal(values.get("amount", ""))
        if (parsed_amount is None or parsed_amount == 0) and not values.get("direction"):
            signed_amounts = list(signed_currency_amount_re.finditer(body))
            if signed_amounts:
                signed_amount = signed_amounts[-1]
                values["amount"] = signed_amount.group("amount")
                values["direction"] = _infer_direction(_parse_decimal(signed_amount.group("amount")), "")
                values["description"] = body[: signed_amount.start()].strip() or values["description"]
        parsed.append(
            _line_from_row(
                len(parsed) + 1,
                values,
                IDENTITY_STATEMENT_MAPPING,
            )
        )
    return tuple(parsed)


def parse_statement_text(text: str) -> tuple[StatementLine, ...]:
    html_grid = _grid_from_html_table(text)
    if html_grid:
        return _lines_from_grid(html_grid)
    delimited = _parse_delimited_statement_text(text)
    if delimited:
        return delimited
    return _parse_loose_statement_text(text)


def _parse_binary_xls(content: bytes) -> tuple[StatementLine, ...]:
    try:
        import xlrd
    except ImportError:
        return ()
    workbook = xlrd.open_workbook(file_contents=content)
    if not workbook.nsheets:
        return ()
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows < 2:
        return ()
    grid = [
        [sheet.cell_value(row_index, column) for column in range(sheet.ncols)]
        for row_index in range(sheet.nrows)
    ]
    return _lines_from_grid(grid)


def _decode_pdf_literal(raw: bytes) -> str:
    text = raw.decode("latin-1", errors="ignore")
    replacements = {
        r"\(": "(",
        r"\)": ")",
        r"\\": "\\",
        r"\n": "\n",
        r"\r": "\n",
        r"\t": "\t",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text


def _fallback_pdf_text(content: bytes) -> str:
    fragments = [
        _decode_pdf_literal(match.group(1))
        for match in re.finditer(rb"\((.*?)\)\s*Tj", content, re.DOTALL)
    ]
    return "\n".join(fragment for fragment in fragments if fragment.strip())


def _extract_pdf_text(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return _fallback_pdf_text(path.read_bytes())
    try:
        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        return _fallback_pdf_text(path.read_bytes())


def parse_statement_csv(path: Path) -> tuple[StatementLine, ...]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return parse_statement_text(handle.read())


def parse_statement_xlsx(path: Path) -> tuple[StatementLine, ...]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        grid = [list(values) for values in sheet.iter_rows(values_only=True)]
        return _lines_from_grid(grid)
    finally:
        workbook.close()


def parse_statement_xls(path: Path) -> tuple[StatementLine, ...]:
    content = path.read_bytes()
    if content.startswith(OLE2_SIGNATURE):
        return _parse_binary_xls(content)
    decoded = _decode_text_export(content)
    return parse_statement_text(decoded) if decoded else ()


def parse_statement_pdf(path: Path) -> tuple[StatementLine, ...]:
    text = _extract_pdf_text(path)
    return parse_statement_text(text) if text.strip() else ()


def parse_statement_file(path: Path) -> tuple[StatementLine, ...]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_statement_csv(path)
    if suffix == ".xls":
        return parse_statement_xls(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_statement_xlsx(path)
    if suffix == ".pdf":
        return parse_statement_pdf(path)
    return ()
